from __future__ import annotations

import csv
import math
import re
from pathlib import Path

import openpyxl
import xlsxwriter


ROOT_DIR = Path(__file__).resolve().parents[1]
OUTPUTS_DIR = ROOT_DIR / "outputs"
VALIDATION_DIR = OUTPUTS_DIR / "ftc_validation"
WORKBOOK_PATH = VALIDATION_DIR / "ftc_validasi_entropy_overlap.xlsx"

CANDIDATE_CSV = OUTPUTS_DIR / "ftc_candidate_table.csv"
SELECTED_CSV = OUTPUTS_DIR / "ftc_selected_clusters.csv"
COMMENTS_CSV = OUTPUTS_DIR / "comment_topics_clusters.csv"


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as file:
        return list(csv.DictReader(file))


def parse_doc_ids(cluster_candidate: str) -> list[int]:
    return [int(match.group(1)) for match in re.finditer(r"D(\d+)", cluster_candidate)]


def excel_col_name(col_idx_zero_based: int) -> str:
    value = col_idx_zero_based + 1
    name = ""
    while value:
        value, remainder = divmod(value - 1, 26)
        name = chr(65 + remainder) + name
    return name


def as_int(value: str) -> int:
    return int(float(value or 0))


def as_float(value: str) -> float:
    return float(value or 0)


def build_workbook() -> Path:
    VALIDATION_DIR.mkdir(parents=True, exist_ok=True)

    candidate_rows = read_csv(CANDIDATE_CSV)
    selected_rows = read_csv(SELECTED_CSV)
    comment_rows = read_csv(COMMENTS_CSV)

    iteration_one = sorted(
        [row for row in candidate_rows if as_int(row["iteration"]) == 1],
        key=lambda row: as_int(row["no_cluster"]),
    )

    doc_count = len(comment_rows)
    candidate_count = len(iteration_one)
    doc_start_col = 5  # F
    doc_end_col = doc_start_col + doc_count - 1
    doc_start_name = excel_col_name(doc_start_col)
    doc_end_name = excel_col_name(doc_end_col)
    candidate_start_excel_row = 8
    candidate_end_excel_row = candidate_start_excel_row + candidate_count - 1

    memberships: list[set[int]] = [
        set(parse_doc_ids(row["cluster_candidate"])) for row in iteration_one
    ]
    doc_frequencies = {
        doc_id: sum(1 for member_docs in memberships if doc_id in member_docs)
        for doc_id in range(1, doc_count + 1)
    }
    doc_contributions = {
        doc_id: (math.log(freq) / freq if freq > 1 else 0.0)
        for doc_id, freq in doc_frequencies.items()
    }
    excel_eo_values = [
        sum(doc_contributions[doc_id] for doc_id in member_docs)
        for member_docs in memberships
    ]

    workbook = xlsxwriter.Workbook(WORKBOOK_PATH)
    workbook.set_calc_mode("auto")

    title_fmt = workbook.add_format(
        {"bold": True, "font_color": "white", "bg_color": "#173B7A", "font_size": 16}
    )
    subtitle_fmt = workbook.add_format({"italic": True, "font_color": "#26364D"})
    header_fmt = workbook.add_format(
        {
            "bold": True,
            "font_color": "white",
            "bg_color": "#173B7A",
            "align": "center",
            "valign": "vcenter",
            "border": 1,
            "text_wrap": True,
        }
    )
    text_fmt = workbook.add_format({"border": 1, "valign": "top", "text_wrap": True})
    number_fmt = workbook.add_format({"border": 1, "num_format": "0.000000"})
    integer_fmt = workbook.add_format({"border": 1, "num_format": "0"})
    formula_fmt = workbook.add_format(
        {"border": 1, "num_format": "0.000000", "bg_color": "#F8FBFF"}
    )
    member_fmt = workbook.add_format({"border": 1, "num_format": "0", "align": "center"})

    # Sheet 1: FTC candidates, close to the lecture-table format.
    ws_candidates = workbook.add_worksheet("FTC Candidates")
    ws_candidates.hide_gridlines(2)
    ws_candidates.merge_range(
        "A1:G1", "Tahap 1. Kemunculan frequent term pada beberapa dokumen", title_fmt
    )
    ws_candidates.merge_range(
        "A2:G2",
        f"FTC Iterasi 1, minimum support = 8 dokumen, jumlah kandidat = {candidate_count}",
        subtitle_fmt,
    )
    ws_candidates.write_row(
        "A4",
        ["No Cluster", "Frequent term set", "Cluster candidate", "Support", "EO Python", "EO Excel", "Selisih"],
        header_fmt,
    )

    for idx, row in enumerate(iteration_one, start=0):
        excel_row = 5 + idx
        validation_row = candidate_start_excel_row + idx
        eo_python = as_float(row["eo"])
        eo_excel_value = excel_eo_values[idx]
        ws_candidates.write_number(excel_row - 1, 0, as_int(row["no_cluster"]), integer_fmt)
        ws_candidates.write(excel_row - 1, 1, row["frequent_term_set"], text_fmt)
        ws_candidates.write(excel_row - 1, 2, row["cluster_candidate"], text_fmt)
        ws_candidates.write_number(excel_row - 1, 3, as_int(row["support"]), integer_fmt)
        ws_candidates.write_number(excel_row - 1, 4, eo_python, number_fmt)
        ws_candidates.write_formula(
            excel_row - 1,
            5,
            f"='EO Validation'!E{validation_row}",
            formula_fmt,
            eo_excel_value,
        )
        ws_candidates.write_formula(
            excel_row - 1,
            6,
            f"=ABS(E{excel_row}-F{excel_row})",
            formula_fmt,
            abs(eo_python - eo_excel_value),
        )

    ws_candidates.set_column("A:A", 12)
    ws_candidates.set_column("B:B", 22)
    ws_candidates.set_column("C:C", 88)
    ws_candidates.set_column("D:G", 13)
    ws_candidates.freeze_panes(4, 0)

    # Sheet 2: formula-driven EO validation.
    ws_validation = workbook.add_worksheet("EO Validation")
    ws_validation.hide_gridlines(2)
    ws_validation.merge_range(
        "A1:E1", "Tahap 2. Penerapan Persamaan Entropy Overlap (EO)", title_fmt
    )
    ws_validation.merge_range(
        "A2:E2",
        "EO dihitung dengan kontribusi per dokumen: ((-1/f_d) * LN(1/f_d)); f_d adalah jumlah kandidat cluster yang memuat dokumen D tersebut.",
        subtitle_fmt,
    )
    ws_validation.write("A4", "Frekuensi dokumen f_d")
    ws_validation.write("A5", "Kontribusi EO per dokumen")

    doc_headers = [f"D{doc_id}" for doc_id in range(1, doc_count + 1)]
    for col_offset, doc_header in enumerate(doc_headers):
        col = doc_start_col + col_offset
        col_name = excel_col_name(col)
        doc_id = col_offset + 1
        ws_validation.write(3, col, doc_header, header_fmt)
        ws_validation.write_formula(
            3,
            col,
            f"=SUM({col_name}${candidate_start_excel_row}:{col_name}${candidate_end_excel_row})",
            integer_fmt,
            doc_frequencies[doc_id],
        )
        ws_validation.write_formula(
            4,
            col,
            f"=IF({col_name}$4>1,((-1/{col_name}$4)*LN(1/{col_name}$4)),0)",
            formula_fmt,
            doc_contributions[doc_id],
        )

    ws_validation.write_row(
        6,
        0,
        ["No Cluster", "Frequent term set", "Cluster candidate", "Support", "EO Excel", *doc_headers],
        header_fmt,
    )

    for idx, (row, member_docs) in enumerate(zip(iteration_one, memberships), start=0):
        excel_row = candidate_start_excel_row + idx
        row_idx = excel_row - 1
        ws_validation.write_number(row_idx, 0, as_int(row["no_cluster"]), integer_fmt)
        ws_validation.write(row_idx, 1, row["frequent_term_set"], text_fmt)
        ws_validation.write(row_idx, 2, row["cluster_candidate"], text_fmt)
        ws_validation.write_formula(
            row_idx,
            3,
            f"=SUM({doc_start_name}{excel_row}:{doc_end_name}{excel_row})",
            formula_fmt,
            len(member_docs),
        )
        ws_validation.write_formula(
            row_idx,
            4,
            f"=SUMPRODUCT({doc_start_name}{excel_row}:{doc_end_name}{excel_row},${doc_start_name}$5:${doc_end_name}$5)",
            formula_fmt,
            excel_eo_values[idx],
        )
        for doc_id in range(1, doc_count + 1):
            ws_validation.write_number(
                row_idx,
                doc_start_col + doc_id - 1,
                1 if doc_id in member_docs else 0,
                member_fmt,
            )

    ws_validation.set_column("A:A", 12)
    ws_validation.set_column("B:B", 22)
    ws_validation.set_column("C:C", 88)
    ws_validation.set_column("D:E", 13)
    ws_validation.set_column(doc_start_col, doc_end_col, 5)
    ws_validation.freeze_panes(7, 5)

    # Sheet 3: selected clusters per FTC iteration.
    ws_selected = workbook.add_worksheet("Selected Clusters")
    ws_selected.hide_gridlines(2)
    ws_selected.merge_range("A1:F1", "Cluster Terpilih FTC per Iterasi", title_fmt)
    ws_selected.write_row(
        "A4",
        ["Iteration", "Cluster ID", "Frequent termset", "Support", "Entropy Overlap", "Remaining docs after"],
        header_fmt,
    )
    for idx, row in enumerate(selected_rows, start=5):
        ws_selected.write_number(idx - 1, 0, as_int(row["iteration"]), integer_fmt)
        ws_selected.write_number(idx - 1, 1, as_int(row["cluster_id"]), integer_fmt)
        ws_selected.write(idx - 1, 2, row["frequent_termset"], text_fmt)
        ws_selected.write_number(idx - 1, 3, as_int(row["support"]), integer_fmt)
        ws_selected.write_number(idx - 1, 4, as_float(row["entropy_overlap"]), number_fmt)
        ws_selected.write_number(idx - 1, 5, as_int(row["n_remaining_docs_after"]), integer_fmt)

    ws_selected.set_column("A:B", 12)
    ws_selected.set_column("C:C", 28)
    ws_selected.set_column("D:F", 18)
    ws_selected.freeze_panes(4, 0)

    workbook.close()
    return WORKBOOK_PATH


def verify_workbook(path: Path) -> None:
    workbook = openpyxl.load_workbook(path, data_only=False, read_only=True)
    expected_sheets = {"FTC Candidates", "EO Validation", "Selected Clusters"}
    missing_sheets = expected_sheets - set(workbook.sheetnames)
    if missing_sheets:
        raise RuntimeError(f"Sheet tidak ditemukan: {sorted(missing_sheets)}")

    validation = workbook["EO Validation"]
    candidates = workbook["FTC Candidates"]

    if validation["E8"].data_type != "f":
        raise RuntimeError("Formula EO Validation!E8 tidak ditemukan.")
    if candidates["F5"].data_type != "f":
        raise RuntimeError("Formula FTC Candidates!F5 tidak ditemukan.")

    workbook.close()


if __name__ == "__main__":
    result_path = build_workbook()
    verify_workbook(result_path)
    print(result_path)
